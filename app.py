# -*- coding: utf-8 -*-
"""
Created on Wed Apr 10 14:06:51 2024

@author: Administrator
"""
from my_module import mendian_format, meituan_caipin_format,format_meituan_table,format_hualala_table,format_zhongtai_table
from flask import Flask, request, url_for, render_template,render_template_string,send_from_directory
import os
import pandas as pd
import numpy as np
import time
import openpyxl
from datetime import datetime, timedelta
import markdown
from dateutil.relativedelta import relativedelta
import socket,requests
import math
import re
import openpyxl
from openpyxl.styles import PatternFill
import glob
#输出文件夹
folder = 'outputs'

def calculate_periods(current_start_date: datetime, current_end_date: datetime):
    days_in_period = (current_end_date - current_start_date).days
    previous_period_end_date = current_start_date - timedelta(days= 1)
    previous_period_start_date = previous_period_end_date - timedelta(days=days_in_period)
    last_year_start_date = current_start_date - relativedelta(years=1)
    last_year_end_date = current_end_date - relativedelta(years=1)

    return {
        '本期': (current_start_date, current_end_date),
        '环比期': (previous_period_start_date, previous_period_end_date),
        '同比期': (last_year_start_date, last_year_end_date)
    }

# 判断销售报货
def sales_status(row):
    col_name_1 = f'{cunhuo_name}报货数量'
    col_name_2 = f'{caipin_name}销售数量'
    if row[col_name_1] == 0:
        baohuo = '无报货'
    else:
        baohuo = '有报货'
    if row[col_name_2] == 0:
        xiaoshou = '无销售'
    else:
        xiaoshou = '有销售'
    return baohuo+xiaoshou

# 判断报货
def check_baohuo(i):
    if i > 0:
        return '有报货'
    else:
        return '无报货'

#计算周期
def calc_zhouqi(days):
    if days == '-':
        zhouqi = '90日内无报货'
    elif days< 30:
        zhouqi = '30日内有报货'
    elif days <60:
        zhouqi = '60日内有报货'
    elif days <90:
        zhouqi = '90日内有报货'
    else:
        zhouqi = '90日内无报货'
    return zhouqi


#计算店内优惠金额(会员日)
def calculate_store_discount(row):
    if row['订单子来源'] == '店内点餐':
        store_discount = row['订单金额(元)'] - row['订单收入(元)']
    else:
        store_discount = 0
    return store_discount

def jiexi(customer_code):
    try:
        filtered_rows = baohuo_df[baohuo_df['客商编码'] == customer_code ]
        max_date = filtered_rows['单据日期'].max()
        result = filtered_rows[filtered_rows['单据日期'] == max_date]
        count = result['数量'].iloc[0]
    except:
        max_date = '-'
        count = '-'
    return max_date, count

# 通过中文逗号分裂，并添加序号
def add_number(text):
    if type(text) == float or len(text) ==0:
        return ""
    text = text.lstrip("，")
    # 根据回车符分割文本
    lines = text.split('，')
    # 给每一行添加编号
    numbered_lines = [f'{i + 1}.{line}' for i, line in enumerate(lines)]
    # 将编号后的行重新组合成文本
    numbered_text = '\n'.join(numbered_lines)
    return numbered_text

# 监控在线状况判断
def cal_zaixian(zaixian):
    a,b = zaixian.split('/')
    if a =='0':
        return '离线'
    elif b=='1':
        return '部分在线'
    elif a==b :
        return '完全在线'
    else :
        return '部分在线'
# 监控表透视
def jiankong_process_stores(df):
    pivot_stores = pd.pivot_table(
        df,
        index=["大区经理", "省区经理", "区域经理"],
        columns=["在线状况"],
        values="门店编号",
        aggfunc="count",
        fill_value=0,
    )
    pivot_stores.reset_index(inplace=True)
    pivot_stores["监控门店数"] = pivot_stores["完全在线"] + pivot_stores["部分在线"] + pivot_stores["离线"]
    return pivot_stores

# 门店周期管理加序号
def add_number(text):
    if type(text) == float or text == '':
        return ""
    # 根据回车符分割文本
    lines = text.split('\n')
    # 给每一行添加编号
    numbered_lines = [f'{i + 1}.{line}' for i, line in enumerate(lines)]
    # 将编号后的行重新组合成文本
    numbered_text = '\n'.join(numbered_lines)
    return numbered_text

# 创建 Flask 应用
app = Flask(__name__, template_folder='./templates',static_folder='outputs')

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.route('/')
def index():
    output_folder = 'Outputs'
    file_list = []
    for file in os.listdir(output_folder):
        if file == 'ReadMe.txt':
            continue
        file_path = os.path.join(output_folder, file)
        creation_date = datetime.fromtimestamp(os.path.getctime(file_path))
        file_list.append((file, creation_date))
    sorted_file_list = sorted(file_list, key=lambda x: x[1], reverse=True)
    if len(sorted_file_list) > 13:
       sorted_file_list = sorted_file_list[:13]
    return render_template('index.html', file_list=sorted_file_list)
# 新品
@app.route('/xinpin')
def xinpin():
    return render_template('xinpin.html')
# 会员日
@app.route('/huiyuan')
def huiyuan():
    return render_template('huiyuan.html')
# 门店格式化
@app.route('/mendian')
def mendian():
    return render_template('mendian.html')
# 监控报表
@app.route('/jiankong')
def jiankong():
    return render_template('jiankong.html')
# 报表格式化
@app.route('/geshihua')
def geshihua():
    return render_template('geshihua.html')

# 文档
@app.route('/readme')
def readme():
    with open('templates/log.md', 'r', encoding='utf-8') as file:
        markdown_text = file.read()
    html_text = markdown.markdown(markdown_text, extensions=['markdown.extensions.tables'])
    return render_template('log.html', content=html_text)


@app.route('/xiaoshou')
def xiaoshou():
    return render_template('xiaoshou.html')

@app.route('/xundian')
def xundian():
    return render_template('xundian.html')

@app.route('/zhouqi')
def zhouqi():
    return render_template('zhouqi.html')

@app.route('/caipin')
def caipin():
    return render_template('caipin.html')

@app.route('/danpin')
def danpin():
    return render_template('danpin.html')

# 定义上传文件的路由
@app.route('/xinpin_upload', methods=['POST'])
def xinpin_upload_files():
    global now
    now = time.strftime('%Y%m%d_%H%M', time.localtime())
    # 获取上传的文件
    file1 = request.files['file1']
    file2 = request.files['file2']
    file3 = request.files.get('file3')
    file1_path = os.path.join('uploads', f'{now}_{file1.filename}')
    file1.save(file1_path)
    file2_path = os.path.join('uploads', f'{now}_{file2.filename}')
    file2.save(file2_path)
    
    if file3:
        file3_path = os.path.join('uploads', f'{now}_{file3.filename}')
        file3.save(file3_path)
        gen_file_name = xinpin_process_files(file1_path, file2_path, file3_path)
    else:
        gen_file_name = xinpin_process_files(file1_path, file2_path)
        
    file_link = url_for('static', filename=gen_file_name)
    html = f'<a href="{file_link}">下载{gen_file_name}</a>'
    return render_template_string(html)


# 销售格式化
@app.route('/geshihua_upload', methods=['POST'])
def geshihua_upload_files():
    global now
    now = time.strftime('%Y%m%d_%H%M', time.localtime())
    # 获取上传的文件
    file1 = request.files['file1']
    file2 = request.files['file2']
    mendian_file_name = f'{now}_门店管理原表.xlsx'
    xiaoshou_file_name = f'{now}_菜品销售原表.xlsx'
    file1_path = os.path.join('uploads', mendian_file_name)
    file1.save(file1_path)
    file2_path = os.path.join('uploads', xiaoshou_file_name)
    file2.save(file2_path)
    mendian_df = mendian_format(file1_path)
    try:
        df_check = pd.read_excel(file2_path, nrows=2)
        if type(df_check.iloc[0][0]) !=float:
            if '按订单来源统计' in df_check.iloc[0][0]:
                pattern = r'\d{4}/\d{2}/\d{2}'
                dates = re.findall(pattern,  df_check.iloc[0][0])
                shiduan = dates[0] + '~' + dates[1]
                df = format_meituan_table(file2_path)
                
            elif '142渠道销售统计表' in df_check.columns[0]:
                pattern = r'(\d{4})(\d{2})(\d{2})--(\d{4})(\d{2})(\d{2})'
                date_range = re.search(pattern, df_check.columns[0])
                start_date = f"{date_range.group(1)}/{date_range.group(2)}/{date_range.group(3)}"
                end_date = f"{date_range.group(4)}/{date_range.group(5)}/{date_range.group(6)}"
                shiduan = f"{start_date}~{end_date}"
                df = format_hualala_table(file2_path)
    except:
        df_check = pd.read_csv(file2_path, nrows=2,encoding='gbk')
        shiduan = df_check['时段'][0]
        shiduan = shiduan.replace('\t','')
        df = format_zhongtai_table(file2_path)

    mendian_df['时段'] = shiduan
    start_date, end_date = shiduan.split("~")
    
    formatted_start_date = start_date.replace("/", "")
    formatted_end_date = end_date.replace("/", "")
    formatted_date_range = f"{formatted_start_date}_{formatted_end_date}"
    
    df_result = pd.merge(mendian_df, df, how='left', left_on='门店编码',right_on = '门店编码')
    df_result = df_result.fillna(0)
    output_filename = f'{folder}\\各渠道销售统计_{formatted_date_range}_{now}.xlsx'
    df_result.to_excel(output_filename,index=False)
    file_link = url_for('static', filename=f'各渠道销售统计_{formatted_date_range}_{now}.xlsx')
    html = f'<a href="{file_link}">下载各渠道销售统计_{formatted_date_range}_{now}.xlsx</a>'
    return render_template_string(html)
    
    

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(folder, filename, as_attachment=True)

# 处理文件
def xinpin_process_files(file1_path, file2_path, file3_path=None):
    global cunhuo_name,caipin_name
    if file3_path is None:
        mendian_format_df = mendian_format(file1_path)
        baohuo_df = pd.read_excel(file2_path, header = 5, converters={'客商编码': str})
        cunhuo_name = baohuo_df['存货名称'].unique()[0]
        col_name = f'{cunhuo_name}报货周期'  # 报货情况
        baohuo_df = baohuo_df.loc[:,['客商编码','数量']]
        baohuo_df = baohuo_df.groupby('客商编码')['数量'].sum()
        df_merge= pd.merge(mendian_format_df, baohuo_df, how='left', left_on='U8C客商编码',right_on='客商编码').fillna(0)
        df_merge.rename(columns={'数量':f'{cunhuo_name}报货数量'}, inplace=True)
        df_merge[col_name] = df_merge[f'{cunhuo_name}报货数量'].apply(check_baohuo)
        df_merge_weibaohuo = df_merge[(df_merge[col_name] == '无报货') & (df_merge['运营状态'] == '营业中')]
        with pd.ExcelWriter(f'{folder}\\{cunhuo_name}_(新品)报货信息_{now}.xlsx') as writer:
            df_merge.to_excel(writer, sheet_name=f'{cunhuo_name}底表', index=False)
            df_merge_weibaohuo.to_excel(writer, sheet_name='未报货表', index=False)
        return f'{cunhuo_name}_(新品)报货信息_{now}.xlsx'
    else:
        mendian_format_df = mendian_format(file1_path)
        baohuo_df = pd.read_excel(file2_path, header = 5, converters={'客商编码': str})
        cunhuo_name = baohuo_df['存货名称'].unique()[0]
        col_name = f'{cunhuo_name}报货周期'  # 报货情况
        baohuo_df = baohuo_df.loc[:,['客商编码','数量']]
        baohuo_df = baohuo_df.groupby('客商编码')['数量'].sum()
        df_merge= pd.merge(mendian_format_df, baohuo_df, how='left', left_on='U8C客商编码',right_on='客商编码').fillna(0)
        df_merge.rename(columns={'数量':f'{cunhuo_name}报货数量'}, inplace=True)
        xiaoshou_df,caipin_name = meituan_caipin_format(file3_path)
        df_merge= pd.merge(df_merge, xiaoshou_df, how='left', left_on='门店编码',right_on='机构编码').fillna(0)
        df_merge.rename(columns={'销售数量':f'{caipin_name}销售数量'}, inplace=True)
        df_merge[col_name] = df_merge.apply(sales_status, axis=1)
        df_merge_weibaohuo = df_merge[(df_merge[col_name] == '无报货无销售') & (df_merge['运营状态'] == '营业中')]
        with pd.ExcelWriter(f'{folder}\\{cunhuo_name}_(新品)销售报货信息_{now}.xlsx') as writer:
            df_merge.to_excel(writer, sheet_name=f'{cunhuo_name}底表', index=False)
            df_merge_weibaohuo.to_excel(writer, sheet_name='未报未销表', index=False)
        
        return f'{food_name}_(新品)销售报货信息_{now}.xlsx'    

@app.route('/danpin_upload', methods=['POST'])
def danpin_upload_files():
    global now
    now = time.strftime('%Y%m%d_%H%M', time.localtime())
    file1 = request.files['file1']
    file2 = request.files['file2']
    mendian_file_name = f'{now}_门店管理原表.xlsx'
    danping_file_name = f'{now}_单品报货原表.xlsx'
    file1_path = os.path.join('uploads',mendian_file_name)
    file1.save(file1_path)
    file2_path = os.path.join('uploads', danping_file_name)
    file2.save(file2_path)
    gen_file_name = danpin_process_files(file1_path, file2_path)
    file_link = url_for('static', filename=gen_file_name)
    html = f'<a href="{file_link}">下载{gen_file_name}</a>'
    return render_template_string(html)


def xundian_process_files(file1_path, file2_path):
    now = time.strftime('%Y%m%d_%H%M', time.localtime())
    current_month = time.strftime("%Y-%m", time.localtime())
    output_filename = f'{folder}\\市场部巡店表整合_{now}.xlsx'
    # 巡店表
    xundian_df = pd.read_excel(file2_path, header = 2)
    xundian_df = xundian_df[xundian_df['巡检日期'].str.contains(current_month)]
    if len(xundian_df) ==0:
        xundian_df = pd.read_excel(file2_path, header = 2)
    xundian_df = xundian_df.sort_values(by='巡检日期', ascending=True)
    xundian_df = xundian_df.fillna("")
    #xundian_df.dropna(subset=['门店编号', '大区经理', '非原物料得分', '过期'], inplace=True)
    xundian_df = xundian_df.dropna(subset=['门店编号', '合计得分', '非原物料得分', '过保质期和变质得分','隔夜物料得分','效期问题得分','营业额同比提升得分','线上平台得分','私域群运营得分',
                                           '交叉污染得分','水质和冰块得分','证照齐全公示且有效得分','现场操作得分','整体形象得分','仪容仪表得分','店内卫生得分','设备设施清洁保养得分','服务态度得分',
                                           ])
    
    # 门店表
    mendian_df = mendian_format(file1_path)
    xundian_df = pd.merge(xundian_df, mendian_df, how='left', left_on='门店编号',right_on='门店编码',suffixes=('_巡店表', ''))
    mendian_df['是否交叉巡店'] = 0
    mendian_df['交叉巡店记录'] = ''
    mendian_df['是否自查'] = 0
    mendian_df['自查记录'] = ''
    # 架构
    jiagou_df = mendian_df[['大区经理','省区经理','区域经理']]
    jiagou_df = jiagou_df.drop_duplicates()
    jiagou_df['自查数量'] = 0
    jiagou_df['跨区数量'] = 0
    
    for index, row in mendian_df.iterrows():
        mdbh = row.门店编码
        qyjl = row.区域经理 #当前区域经理
        xd_df = xundian_df[xundian_df['门店编号'] == mdbh] #巡店记录
        if len(xd_df) == 0 :
             pass # 无此店巡店记录
        else:
            for i,j in xd_df.iterrows():
                xjr = j.巡检人
                xjrq = j.巡检日期
                date_obj = datetime.strptime(xjrq, "%Y-%m-%d %H:%M")
                xjrq = date_obj.strftime("%Y%m%d")
                if xjr == qyjl:#自查
                    mendian_df.loc[index, '是否自查'] = 1
                    mendian_df.loc[index, '自查记录'] = str(mendian_df.loc[index, '自查记录'] ) + xjrq + ':' + xjr
                    jiagou_df.loc[jiagou_df['区域经理'] == xjr, '自查数量']  = jiagou_df.loc[jiagou_df['区域经理'] == xjr, '自查数量']  + 1
                else: #交叉巡店
                    mendian_df.loc[index, '是否交叉巡店'] = 1
                    mendian_df.loc[index, '交叉巡店记录'] = str(mendian_df.loc[index, '交叉巡店记录'] ) + xjrq + ':' + xjr
                    jiagou_df.loc[jiagou_df['区域经理'] == xjr, '跨区数量']  = jiagou_df.loc[jiagou_df['区域经理'] == xjr, '跨区数量']  + 1

    mendian_df['自查记录'] = mendian_df['自查记录'].apply(add_number)
    mendian_df['交叉巡店记录'] = mendian_df['交叉巡店记录'] .apply(add_number)
    mendian_df['是否被巡查'] = np.where((mendian_df['是否交叉巡店'] + mendian_df['是否自查']) == 0, 0, 1)
    mendian_df['门店数量'] = 1 
    filtered_df = mendian_df[mendian_df['运营状态'] == '营业中']
    pivot_df = filtered_df.pivot_table(index=['大区经理', '省区经理', '区域经理'],
                                    values=['门店数量', '是否被巡查', '是否交叉巡店', '是否自查'],
                                    aggfunc='sum')
    pivot_df = pivot_df.reset_index()
    pivot_df['被巡查占比'] = pivot_df['是否被巡查']/pivot_df['门店数量']
    pivot_df['被交叉巡查占比'] = pivot_df['是否交叉巡店']/pivot_df['门店数量']
    pivot_df['自查占比'] = pivot_df['是否自查']/pivot_df['门店数量']
    pivot_df.rename(columns={
        '是否交叉巡店': '被交叉巡店数量',
        '是否自查':'自查数量',
        '是否被巡查':'被巡查数量'
        }, inplace=True)
    pivot_df = pivot_df.loc[:, ['大区经理', '省区经理', '区域经理', '门店数量','被巡查数量','被巡查占比','自查数量','自查占比','被交叉巡店数量','被交叉巡查占比']]
    pivot_df['距40%相差门店数'] = (pivot_df['门店数量'] * 0.4 - pivot_df['被交叉巡店数量']).apply(lambda x: math.ceil(x)).astype(int)
    pivot_df['距40%相差门店数'] = pivot_df['距40%相差门店数'].clip(lower=0, upper=10000)
    # 增加汇总行
    result = pivot_df.pivot_table(index=["大区经理", "省区经理", "区域经理"], aggfunc="sum")
    summary_by_daqu_manager = result.groupby(level="大区经理").sum().reset_index()
    summary_by_daqu_manager['被巡查占比'] = summary_by_daqu_manager['被巡查数量']/summary_by_daqu_manager['门店数量']
    summary_by_daqu_manager['自查占比'] = summary_by_daqu_manager['自查数量']/summary_by_daqu_manager['门店数量']
    summary_by_daqu_manager['被交叉巡查占比'] = summary_by_daqu_manager['被交叉巡店数量']/summary_by_daqu_manager['门店数量']
    
    summary_by_sheng_manager = result.groupby(level=["大区经理", "省区经理"]).sum().reset_index()
    summary_by_sheng_manager['被巡查占比'] = summary_by_sheng_manager['被巡查数量']/summary_by_sheng_manager['门店数量']
    summary_by_sheng_manager['自查占比'] = summary_by_sheng_manager['自查数量']/summary_by_sheng_manager['门店数量']
    summary_by_sheng_manager['被交叉巡查占比'] = summary_by_sheng_manager['被交叉巡店数量']/summary_by_sheng_manager['门店数量']
    
    pivot_df = pd.concat([pivot_df, summary_by_daqu_manager], axis=0, ignore_index=True)
    pivot_df = pd.concat([pivot_df, summary_by_sheng_manager], axis=0, ignore_index=True)
    pivot_df['大区经理'] = pivot_df['大区经理'].astype(manager_type)
    pivot_df = pivot_df.sort_values(['大区经理', '省区经理', '区域经理'], ascending=True)
    
    # 查找“区域经理”列的空值并根据“省区经理”列是否为空进行修改
    pivot_df["区域经理"] = pivot_df.apply(
        lambda row: "大区合计" if pd.isna(row["区域经理"]) and pd.isna(row["省区经理"]) else "省区合计" if pd.isna(row["区域经理"]) else row["区域经理"],
        axis=1,
    )
    #保存
    writer = pd.ExcelWriter(output_filename, engine='openpyxl')
    mendian_df.to_excel(writer, sheet_name='门店被巡详表', index=False)
    pivot_df.to_excel(writer, sheet_name='营业中门店被巡', index=False)
    jiagou_df.to_excel(writer, sheet_name='经理巡店次数', index=False)
    writer.close()
    #设置格式
    sheng_fill_color = 'EEECE1'
    daqu_fill_color = '948A54'
    workbook = openpyxl.load_workbook(output_filename) 
    for worksheet in workbook.worksheets:
        for column in worksheet.columns:
            if '占比' in column[0].value:
                for cell in column:
                    cell.number_format = '0.00%'
        for row in worksheet.iter_rows():
            if row[2].value == "省区合计":
                for cell in row[0:18]: 
                    cell.fill = PatternFill(start_color=sheng_fill_color, end_color=sheng_fill_color, fill_type="solid")
            elif row[2].value == "大区合计":
                for cell in row[0:18]: 
                    cell.fill = PatternFill(start_color=daqu_fill_color, end_color=daqu_fill_color, fill_type="solid")
    workbook.save(output_filename)
    return f'市场部巡店表整合_{now}.xlsx'

@app.route('/xundian_upload', methods=['POST'])
def xundian_upload_files():
    global now
    now = time.strftime('%Y%m%d_%H%M', time.localtime())
    file1 = request.files['file1']
    file2 = request.files['file2']
    mendian_file_name = f'{now}_门店管理原表.xlsx'
    xunjian_file_name = f'{now}_门店综合巡检原表.xlsx'
    file1_path = os.path.join('uploads', mendian_file_name)
    file1.save(file1_path)
    file2_path = os.path.join('uploads', xunjian_file_name)
    file2.save(file2_path)
    gen_file_name = xundian_process_files(file1_path, file2_path)
    file_link = url_for('static', filename=gen_file_name)
    html = f'<a href="{file_link}">下载{gen_file_name}</a>'
    return render_template_string(html)
 

# 处理文件
def danpin_process_files(file1_path, file2_path):
    global food_name,baohuo_df
    mendian_df = mendian_format(file1_path)
    baohuo_df = pd.read_excel(file2_path, header = 5, converters={'客商编码': str})
    try:
        baohuo_df['客商编码'] = baohuo_df['客商编码'].astype(int).astype(str)
    except:
        baohuo_df['客商编码'] = baohuo_df['客商编码'].astype(str) 
    food_name = baohuo_df['存货名称'].loc[0]
    output_filename = f'{folder}\\{food_name}_报货信息_{now}.xlsx'
    mendian_df[['日期', f'{food_name}数量']] = mendian_df['U8C客商编码'].apply(jiexi).apply(pd.Series)
    mendian_df['日期'] = pd.to_datetime(mendian_df['日期'], errors='coerce')
    today_ = datetime.now().date()
    mendian_df['上次报货距今'] = mendian_df['日期'].apply(lambda x: (today_ - x.date()).days if pd.notnull(x) else "-")
    mendian_df[f'{food_name}报货周期'] = mendian_df['上次报货距今'].apply(calc_zhouqi)
    mendian_df.rename(columns={'日期': '上次报货日期'}, inplace=True)
    mendian_df['大区经理'] = mendian_df['大区经理'].astype(manager_type)
    mendian_df = mendian_df.sort_values(by=['大区经理', '省区经理', '区域经理'], ascending=True) # 升序排序
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        sheet_name = f'{food_name}底表'
        mendian_df.to_excel(writer, sheet_name=sheet_name, index=False)
    workbook = openpyxl.load_workbook(output_filename)
    worksheet = workbook.active
    # 设置 L 列的日期格式
    for row in range(1, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=12)  # L 列的索引为 12
        cell.number_format = 'yyyy/m/d'
    workbook.save(output_filename)
    return f'{food_name}_报货信息_{now}.xlsx'  


@app.route('/caipin_upload', methods=['POST'])
def caipin_upload_files():
    now = time.strftime('%Y%m%d_%H%M', time.localtime())
    output_filename = f'{folder}\\菜品标准名称还原_{now}.xlsx'
    # 获取上传的文件
    file1 = request.files['file1']
    file2 = request.files['file2']
    mendian_file_name = f'{now}_门店管理原表.xlsx'
    caipin_file_name = f'{now}_菜品销售原表.xlsx'
    file1_path = os.path.join('uploads',mendian_file_name)
    file1.save(file1_path)
    file2_path = os.path.join('uploads', caipin_file_name)
    file2.save(file2_path)
    try:
        caipin_df = pd.read_excel(file1_path,header=2)
        caipin_df = caipin_df.drop(caipin_df.iloc[0:2].index)
        caipin_df.loc[:,['菜品名称']]
    except:
        caipin_df = pd.read_excel(file1_path)
    duizhao_df = pd.read_excel(file2_path,header=2)
    duizhao_df = duizhao_df[duizhao_df['流程状态_系统字段'] == '已结束']
    duizhao_df = duizhao_df.loc[:,['新增套餐名单品名','标准单品名称','是否为套餐','数量']]
    result_df = pd.merge(caipin_df, duizhao_df, how='left', left_on='菜品名称',right_on='新增套餐名单品名')
    try:
        result_df['合计数量'] = result_df['销售数量'] *result_df['数量']
    except:
        pass
    result_df.to_excel(output_filename,sheet_name='菜品名称还原', index=False)
    file_link = url_for('static', filename=f'菜品标准名称还原_{now}.xlsx')
    html = f'<a href="{file_link}">下载菜品标准名称还原_{now}.xlsx</a>'
    return html


@app.route('/xiaoshou_upload', methods=['POST'])
def xiaoshou_upload_files():
    now = time.strftime('%Y%m%d_%H%M', time.localtime())
    file1 = request.files['file1']
    file2 = request.files['file2']
    mendian_file_name = f'{now}_门店管理原表.xlsx'
    caipin_file_name = f'{now}_菜品销售原表.xlsx'
    file1_path = os.path.join('uploads', mendian_file_name)
    file1.save(file1_path)
    file2_path = os.path.join('uploads',caipin_file_name)
    file2.save(file2_path)
    mendian_df = mendian_format(file1_path)
    xiaoshou_df,caipin_name = meituan_caipin_format(file2_path)
    df_xiaoshou = pd.merge(mendian_df, xiaoshou_df, how='left', left_on='门店编码',right_on='机构编码').fillna(0)
    col_name = f'{caipin_name}销售'
    df_xiaoshou[col_name] =  df_xiaoshou['销售数量'].apply(lambda x: '有销售' if x > 0 else '无销售')
    output_filename = f'{folder}\\{caipin_name}_单店单品销售统计_{now}.xlsx'
    df_xiaoshou.to_excel(output_filename,sheet_name = '销售信息', index=False)
    file_link = url_for('static', filename=f'{caipin_name}_单店单品销售统计_{now}.xlsx')
    html = f'<a href="{file_link}">下载{caipin_name}_单店单品销售统计_{now}.xlsx</a>'
    return html

@app.route('/mendian_upload', methods=['POST'])
def mendian_upload_files():
    now = time.strftime('%Y%m%d_%H%M', time.localtime())
    output_filename = f'{folder}\\门店管理表_{now}.xlsx'
    mendian_file_name = f'{now}_门店管理原表.xlsx'
    file1 = request.files['file1']
    file1_path = os.path.join('uploads', mendian_file_name)
    file1.save(file1_path)
    mendian_df = mendian_format(file1_path)
    mendian_df.to_excel(output_filename, sheet_name='门店信息表',index=False)
    file_link = url_for('static', filename=f'门店管理表_{now}.xlsx')
    html = f'<a href="{file_link}">下载门店管理表_{now}.xlsx</a>'
    return html



# 会员日数据处理
@app.route('/huiyuan_upload', methods=['POST'])
def huiyuan_upload_files():
    now = time.strftime('%Y%m%d_%H%M', time.localtime())
    file1 = request.files['file1']
    caipin_file_name = f'{now}_菜品销售原表.xlsx'
    file1_path = os.path.join('uploads',caipin_file_name)
    file1.save(file1_path)
    try:# 选择了优惠/全选
        df =  pd.read_excel(file1_path,header=[2,3,4],skipfooter=1)
        df.columns = df.columns.map(''.join).str.replace(' ', '')
        old_header = df.columns
        new_header = [s.split('Unnamed:')[0] if 'Unnamed:' in s else s for s in old_header]
        df.columns = new_header
    except:# 未选择优惠
        df =  pd.read_excel(file1_path,header=[2],skipfooter=1)
    df_xiaoshou = df.loc[:,['机构编码','商户号','门店','菜品名称','订单编号','销售数量','订单分类','菜品下单来源','订单子来源','订单金额(元)','订单收入(元)']]
    df_xiaoshou['店内优惠'] = df_xiaoshou.apply(lambda row: row['订单金额(元)'] - row['订单收入(元)'] if row['订单子来源'] == '店内点餐' else 0, axis=1)
    caipin_name = df_xiaoshou['菜品名称'].iloc[0]
    if '冰淇淋' in caipin_name:
        caipin_name = '冰淇淋'
    elif '柠檬水' in caipin_name:
        caipin_name = '柠檬水'
    df_xiaoshou['店内优惠'] = df_xiaoshou.apply(calculate_store_discount,axis=1)
    df_xiaoshou_youhui  = df_xiaoshou[df_xiaoshou['店内优惠'] >= 2] 
    df_xiaoshou['机构编码'] = df_xiaoshou['机构编码'].str.split('-').str[0]
    pivot_all = pd.pivot_table(df_xiaoshou, index='机构编码', values='销售数量', aggfunc='sum').reset_index()
    pivot_youhui = pd.pivot_table(df_xiaoshou_youhui, index='机构编码', values='销售数量', aggfunc='sum').reset_index()
    df_merge= pd.merge(pivot_all, pivot_youhui, how='left', on='机构编码',suffixes=('_总', '_优惠')).fillna(0)
    output_filename = f'{folder}\\{caipin_name}会员日销售情况_{now}.xlsx'
    df_xiaoshou['订单编号'] = df_xiaoshou['订单编号'].astype(str)
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df_xiaoshou.to_excel(writer, sheet_name='底表', index=False)
        df_merge.to_excel(writer, sheet_name='分析表', index=False)
    file_link = url_for('static', filename=f'{caipin_name}会员日销售情况_{now}.xlsx')
    html = f'<a href="{file_link}">{caipin_name}会员日销售情况_{now}.xlsx</a>'
    return html

@app.route('/jiankong_upload', methods=['POST'])
def jiankong_upload_files():
    now = time.strftime('%Y%m%d_%H%M', time.localtime())
    file1 = request.files['file1']
    file2 = request.files['file2']
    mendian_file_name = f'{now}_门店管理原表.xlsx'
    jiankong_file_name = f'{now}_监控状态统计原表.xlsx'
    file1_path = os.path.join('uploads',mendian_file_name)
    file1.save(file1_path)
    file2_path = os.path.join('uploads',jiankong_file_name)
    file2.save(file2_path)
    mendian_df = mendian_format(file1_path)
    jiankong_df = pd.read_excel(file2_path,header=2)
    #去除门店
    quchu_df = pd.read_excel(r"C:\Users\Administrator\OneDrive\甜啦啦\代码底表\监控去除门店.xlsx")
    quchu_store_ids = quchu_df['门店编号'].to_list()
    jiankong_df = jiankong_df[~jiankong_df['门店编号'].isin(quchu_store_ids)]
    #补齐门店编号
    new_store_codes = {'城中万达广场': 'TLL07669'}
    for store_name, new_code in new_store_codes.items():
        jiankong_df.loc[jiankong_df['门店名称'] == store_name, '门店编号'] = new_code
    df_merge= pd.merge(jiankong_df, mendian_df, how='left', left_on ='门店编号',right_on='门店编码',suffixes=('', '_OA'))
    df_merge['在线状况'] = df_merge['设备在线情况'].apply(cal_zaixian)
    df_merge = df_merge.loc[:,['门店编号','门店名称','省','市','区','运营状态','南北战区','大区经理','省区经理','区域经理','设备在线情况','设备在线率','设备存储情况','在线状况']]
    df_open_stores = df_merge.query('运营状态 != "空合同" and 运营状态 != "长期闭店"') # 排除空合同、长期闭店门店
    pivot_open_stores = jiankong_process_stores(df_open_stores)
    df_operating_stores = df_merge[df_merge["运营状态"] == "营业中"] # 营业中门店
    pivot_operating_stores = jiankong_process_stores(df_operating_stores)
    df_result= pd.merge(pivot_open_stores, pivot_operating_stores, how='left', on ='区域经理',suffixes=('', '(营业中)'))
    df_result['在线'] =  df_result['完全在线'] + df_result['部分在线'] 
    df_result['在线(营业中)'] =  df_result['完全在线(营业中)'] + df_result['部分在线(营业中)'] 
    df_result = df_result.loc[:,['大区经理','省区经理','区域经理','监控门店数','在线','完全在线','部分在线','离线','监控门店数(营业中)','在线(营业中)','完全在线(营业中)','部分在线(营业中)','离线(营业中)']]

    output_filename = f'{folder}\\门店监控状态统计_{now}.xlsx'
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df_merge.to_excel(writer, sheet_name='底表', index=False)
        df_result.to_excel(writer, sheet_name='中间表', index=False)
    
    file_link = url_for('static', filename=f'门店监控状态统计_{now}.xlsx')
    html = f'<a href="{file_link}">下载门店监控状态统计_{now}.xlsx</a>'
    return html

@app.route('/zhouqi_upload', methods=['POST'])
def zhouqi_upload_files():
    now = time.strftime('%Y%m%d_%H%M', time.localtime())
    file1 = request.files['file1']
    file2 = request.files['file2']
    mendian_file_name = f'{now}_门店管理原表.xlsx'
    jiankong_file_name = f'{now}_正式合同原表.xlsx'
    file1_path = os.path.join('uploads', mendian_file_name)
    file1.save(file1_path)
    file2_path = os.path.join('uploads', jiankong_file_name)
    file2.save(file2_path)
    md_df = mendian_format(file1_path)
    md_df.set_index('门店编码', inplace=True)
    md_df['迁址日期'] = ''
    md_df['最后一次迁址合同登记日期']= ''
    md_df['过户日期'] = ''
    md_df['最后一次过户合同登记日期'] = ''
    md_df['最后一次续约合同登记日期'] = ''
    
    ht_df = pd.read_excel(file2,header=1)
    ht_df = ht_df.dropna(subset=['合同开始日期', '合同结束日期'])
    custom_order = ['作废','解约','过期','生效']
    ht_df['合同状态'] = ht_df['合同状态'].astype(pd.CategoricalDtype(categories=custom_order, ordered=True))
    ht_df = ht_df.sort_values(by=['登记日期', '合同状态'], ascending=True)
    
    for index, row in md_df.iterrows(): #门店编号 index
        h_df = ht_df[ht_df['门店编号'] == index]
        md_df.loc[index, '合同结束日期'] = h_df['合同结束日期'].max()
        md_df.loc[index, '合同开始日期'] = h_df['合同开始日期']. min()
        for i,j in h_df.iterrows():
            if j.合同类型 == "初签":
                djrq = j.登记日期
                md_df.loc[index, '合同登记日期'] = djrq
            elif j.合同类型 == "补录":
                djrq = j.登记日期
                md_df.loc[index, '合同登记日期'] = str(djrq)+'(补录)'
            elif j.合同类型 == "迁址":
                djrq = j.登记日期
                md_df.loc[index, '迁址日期'] =str(md_df.loc[index, '迁址日期']) +'\n'+ djrq
                md_df.loc[index, '最后一次迁址合同登记日期'] = djrq
            elif j.合同类型 == "过户":
                djrq = j.登记日期
                md_df.loc[index, '过户日期'] =str(md_df.loc[index, '过户日期']) +'\n'+ djrq
                md_df.loc[index, '最后一次过户合同登记日期'] = djrq
            elif j.合同类型 == "续约":
                djrq = j.登记日期
                md_df.loc[index, '最后一次续约合同登记日期'] = djrq
    
    md_df['迁址日期'] = md_df['迁址日期'].str.lstrip('\n')
    md_df['过户日期'] = md_df['过户日期'].str.lstrip('\n')
    
    md_df['迁址日期'] = md_df['迁址日期'].apply(add_number)
    md_df['过户日期'] = md_df['过户日期'].apply(add_number)
    md_df = md_df.reset_index()
    md_df = md_df.loc[:,['门店编码','门店名称','运营状态','大区经理','省区经理','区域经理','合同登记日期','合同开始日期','合同结束日期','迁址日期','最后一次迁址合同登记日期','过户日期','最后一次过户合同登记日期','最后一次续约合同登记日期']]
    md_df = md_df.rename(columns={'最后一次迁址合同登记日期': '最新迁址日期', '最后一次过户合同登记日期': '最新过户日期','最后一次续约合同登记日期':'最新续约日期'})
    
    md_df['合同登记日期'] = md_df['合同登记日期'].replace('nan', '')
    output_filename = f'{folder}\\门店周期管理_{now}.xlsx'
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        md_df.to_excel(writer, sheet_name='门店周期管理', index=False)
        
    file_link = url_for('static', filename=f'门店周期管理_{now}.xlsx')
    html = f'<a href="{file_link}">下载门店周期管理_{now}.xlsx.xlsx</a>'
    return html



@app.route('/qishu', methods=['GET', 'POST'])
def qishu():
    result = []
    if request.method == 'POST':
        start_date_str = request.form['start_date']
        end_date_str = request.form['end_date']
        current_start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        current_end_date = datetime.strptime(end_date_str, '%Y-%m-%d')

        periods = calculate_periods(current_start_date, current_end_date)

        result = []
        for period_name, (start_date, end_date) in periods.items():
            result.append((period_name, start_date, end_date))    
    return render_template('qishu.html', result=result)

def get_local_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('10.255.255.255', 1))
        ip = s.getsockname()[0]
    except Exception:
        ip = '127.0.0.1'
    finally:
        s.close()
    return ip


manager_order = ['严鹤','刘成龙','杨硕','王枫涛','刘波','胡冰雪','赵磊','邵陈龙']
manager_type = pd.CategoricalDtype(categories=manager_order, ordered=True)


post_url = 'https://note.bizha.top/tianlala' 
local_ip = get_local_ip()
post_data = {'text':f'{local_ip}:5000'}
requests.post(post_url,post_data)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000,debug=1)
    