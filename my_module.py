import re
import pandas as pd
import os
import shutil
import glob
import openpyxl
from openpyxl.styles import PatternFill

# 返回所有加盟店
def mendian_format(file):
    df = pd.read_excel(file, sheet_name = '门店信息表', header=1, converters={'U8C客商编码': str})
    df = df[df['大区经理'].notna()] #排除大区经理为空的门店
    df = df[~df['门店名称'].str.contains('测试|茶语|茶讯')] #排除测试门店、教学门店
    df['U8C客商编码'] = df['U8C客商编码'].astype(str)
    # 删除各级经理列中的括号及括号内的内容
    df['大区经理'] = df['大区经理'].str.replace('(\(.*?\))', '', regex=True)
    df['省经理'] = df['省经理'].str.replace('(\(.*?\))', '', regex=True)
    df['区域经理'] = df['区域经理'].str.replace('(\(.*?\))', '', regex=True)
    # 将 "区域经理" 列为空的值填充为 "省经理" 的值
    df['区域经理'] = df['区域经理'].fillna(df['省经理'])
    df = df.loc[:, ['门店编号', '门店名称', '大区经理', '省经理', '区域经理', '南北战区', '运营状态', '省', '市', '区', 'U8C客商编码']]
    df.rename(columns={'门店编号': '门店编码', '省经理': '省区经理'}, inplace=True)
    df = df.sort_values(['南北战区','大区经理', '省区经理', '区域经理'], ascending=True)
    return df

# 返回所有门店（加盟店+直营店）
def all_mendian_format(file):
    df = pd.read_excel(file, sheet_name = '门店信息表', header=1, converters={'U8C客商编码': str})
    df = df[~df['门店名称'].str.contains('测试|茶语|茶讯')] #排除测试门店、教学门店
    df = df[~df['门店编号'].str.contains('KXD')] #排除卡小逗
    df['U8C客商编码'] = df['U8C客商编码'].astype(str)
    # 删除各级经理列中的括号及括号内的内容
    df['大区经理'] = df['大区经理'].str.replace('(\(.*?\))', '', regex=True)
    df['省经理'] = df['省经理'].str.replace('(\(.*?\))', '', regex=True)
    df['区域经理'] = df['区域经理'].str.replace('(\(.*?\))', '', regex=True)
    # 将 "区域经理" 列为空的值填充为 "省经理" 的值
    df['区域经理'] = df['区域经理'].fillna(df['省经理'])
    df = df.loc[:, ['门店编号', '门店名称', '大区经理', '省经理', '区域经理', '南北战区', '运营状态', '省', '市', '区', 'U8C客商编码']]
    df.rename(columns={'门店编号': '门店编码', '省经理': '省区经理'}, inplace=True)
    df = df.sort_values(['南北战区','大区经理', '省区经理', '区域经理'], ascending=True)
    return df

# 美团格式化
def meituan_caipin_format(file):
    df =  pd.read_excel(file,header=[2,3],skipfooter=1)
    df.columns = df.columns.map(''.join).str.replace(' ', '')
    old_header = df.columns
    new_header = [s.split('Unnamed:')[0] if 'Unnamed:' in s else s for s in old_header]
    df.columns = new_header
    caipin_name = df['菜品名称'].unique()[0]
    df['机构编码'] = df['机构编码'].str.split('-').str[0]
    df = df.groupby('机构编码')['销售数量'].sum()
    return df,caipin_name

# 移动文件
def move_file(source_file, target_folder):
    file_name = os.path.basename(source_file)
    target_file = os.path.join(target_folder, file_name)
    shutil.move(source_file, target_file)

# 获取文件夹中所有.xlxs、xls文件
def list_excel_files(folder):
    file_list = glob.glob(os.path.join(folder, '*.xlsx')) + glob.glob(os.path.join(folder, '*.xls'))
    for file in file_list:
        print(file)
    return file_list

#u8c导出文件完整性校验
def check_u8c_export(df):
    second_last_row = df.iloc[-2]
    if second_last_row.存货名称 == '--------' :
        print('数据已完整导出！')
    else: 
        print('数据导出不全！')
    df = df[:-2]
    print('当报货表中，共有%s件商品'%len(df['存货名称'].unique()))
    return df

def format_meituan_table(file):
    columns = ['门店编码', '营业天数','流水金额', '实收金额', '订单数', '堂食流水', '堂食实收', '堂食订单数', '外卖流水', '外卖实收', '外卖订单数', '小程序流水', '小程序实收', '小程序订单数']
    df_emt = pd.DataFrame(columns=columns)
    df =  pd.read_excel(file,header=[2,3,4],skipfooter=1)
    df.columns = df.columns.map(''.join).str.replace(' ', '')
    old_header = df.columns
    new_header = [s.split('Unnamed:')[0] if 'Unnamed:' in s else s for s in old_header]
    df.columns = new_header
    df_emt['门店编码'] = df['机构编码']
    df_emt['营业天数'] = df['营业天数']
    df_emt['流水金额'] = df['营业额（元）']
    df_emt['实收金额'] = df['营业收入（元）']
    df_emt['订单数'] = df['订单量']
    df_emt['堂食流水'] = df['渠道营业构成收银POS营业额（元）']
    df_emt['堂食实收'] = df['渠道营业构成收银POS营业收入（元）']
    df_emt['堂食订单数'] = df['渠道营业构成收银POS订单量']
    df_emt['外卖流水'] = df['渠道营业构成饿了么外卖营业额（元）'] + df['渠道营业构成美团外卖营业额（元）']
    df_emt['外卖实收'] = df['渠道营业构成饿了么外卖营业收入（元）'] + df['渠道营业构成美团外卖营业收入（元）']
    df_emt['外卖订单数'] = df['渠道营业构成饿了么外卖订单量'] + df['渠道营业构成美团外卖订单量']
    df_emt['小程序流水'] = df['渠道营业构成第三方小程序营业额（元）']
    df_emt['小程序实收'] = df['渠道营业构成第三方小程序营业收入（元）']
    df_emt['小程序订单数'] = df['渠道营业构成第三方小程序订单量'] 
    df_emt['门店编码'] = df_emt['门店编码'].str.split('-').str[0]
    pivot_df = df_emt.groupby('门店编码').sum().reset_index()
    pivot_df = pivot_df[columns]
    return pivot_df


def format_hualala_table(file):
    shouyinji_path = r"C:\Users\Administrator\OneDrive\甜啦啦\代码底表\收银机管理表_2024.3.22.xlsx"
    df_syj = pd.read_excel(shouyinji_path)
    df_syj = df_syj[['门店编码','组织编码']]
    df_syj = df_syj.dropna(subset=['组织编码'])
    columns = ['门店编码', '营业天数','流水金额', '实收金额', '订单数', '堂食流水', '堂食实收', '堂食订单数', '外卖流水', '外卖实收', '外卖订单数', '小程序流水', '小程序实收', '小程序订单数']
    df_emt = pd.DataFrame(columns=columns)
    df = pd.read_excel(file,header=[2,3],skipfooter=1)
    df.columns = df.columns.map(''.join).str.replace(' ', '')
    old_header = df.columns
    new_header = [s.split('Unnamed:')[0] if 'Unnamed:' in s else s for s in old_header]
    df.columns = new_header
    df = df.dropna(subset=['店铺组织编码'])
    df['店铺组织编码'] = df['店铺组织编码'].astype(str)
    df= pd.merge(df, df_syj, how='left', left_on='店铺组织编码',right_on = '组织编码')
    df_emt['门店编码'] = df['门店编码']
    df_emt['流水金额'] = df['合计流水金额']
    df_emt['实收金额'] = df['合计实收金额']
    df_emt['订单数'] = df['合计账单数']
    df_emt['外卖流水'] = df['美团外卖流水金额'] + df['饿了么外卖流水金额']
    df_emt['外卖实收'] = df['美团外卖实收金额'] + df['饿了么外卖实收金额']
    df_emt['外卖订单数'] = df['美团外卖账单数'] + df['饿了么外卖账单数']
    df_emt['小程序流水'] = df['微信小程序流水金额'] + df['支付宝小程序流水金额']
    df_emt['小程序实收'] = df['微信小程序实收金额'] + df['支付宝小程序实收金额']
    df_emt['小程序订单数'] = df['微信小程序账单数'] + df['支付宝小程序账单数']
    df_emt['堂食流水'] = df_emt['流水金额'] - df_emt['外卖流水'] - df_emt['小程序流水'] 
    df_emt['堂食实收'] = df_emt['实收金额'] - df_emt['外卖实收'] - df_emt['小程序实收'] 
    df_emt['堂食订单数'] = df_emt['订单数'] - df_emt['外卖订单数'] - df_emt['小程序订单数'] 
    pivot_df = df_emt.groupby('门店编码').sum().reset_index()
    pivot_df = pivot_df[columns]
    pivot_df= pivot_df.fillna(0)
    pivot_df['营业天数'] = ''
    return pivot_df


def format_zhongtai_table(file):
    df = pd.read_csv(file, encoding='gbk')
    df = df.replace('\t', '', regex=True)
    #日透视
    pivot_df_day = pd.pivot_table(df, index=['门店编码','日期'], columns='渠道', values=['流水金额', '实收金额', '订单数'], aggfunc='sum')
    pivot_df_day.columns = pivot_df_day.columns.map('_'.join)
    pivot_df_day = pivot_df_day.fillna(0)
    pivot_df_day = pivot_df_day.reset_index()
    pivot_df_day['流水金额'] = pivot_df_day['流水金额_pos'] + pivot_df_day['流水金额_小程序'] + pivot_df_day['流水金额_美团'] + pivot_df_day['流水金额_饿了么']
    pivot_df_day['实收金额'] = pivot_df_day['实收金额_pos'] + pivot_df_day['实收金额_小程序'] + pivot_df_day['实收金额_美团'] + pivot_df_day['实收金额_饿了么']
    pivot_df_day['订单数'] = pivot_df_day['订单数_pos'] + pivot_df_day['订单数_小程序'] + pivot_df_day['订单数_美团'] + pivot_df_day['订单数_饿了么']
    pivot_df_day['外卖流水'] = pivot_df_day['流水金额_美团'] + pivot_df_day['流水金额_饿了么']
    pivot_df_day['外卖实收'] = pivot_df_day['实收金额_美团'] + pivot_df_day['实收金额_饿了么']
    pivot_df_day['外卖订单数'] = pivot_df_day['订单数_美团'] + pivot_df_day['订单数_饿了么']
    pivot_df_day['营业天数'] = pivot_df_day['流水金额'].apply(lambda x: 1 if x > 0 else 0)
    pivot_df_day = pivot_df_day.rename(columns={
        '流水金额_pos': '堂食流水',
        '实收金额_pos': '堂食实收',
        '订单数_pos':'堂食订单数',
        '流水金额_小程序':'小程序流水',
        '实收金额_小程序':'小程序实收',
        '订单数_小程序':'小程序订单数'
    })
    pivot_df_day = pivot_df_day.loc[:,['门店编码','日期','营业天数','流水金额','实收金额','订单数','堂食流水','堂食实收','堂食订单数','外卖流水','外卖实收','外卖订单数','小程序流水','小程序实收','小程序订单数']]
    pivot_result = pd.pivot_table(pivot_df_day, index='门店编码',  aggfunc='sum').reset_index()
    pivot_result = pivot_result.loc[:,['门店编码','营业天数','流水金额','实收金额','订单数','堂食流水','堂食实收','堂食订单数','外卖流水','外卖实收','外卖订单数','小程序流水','小程序实收','小程序订单数']]
    return pivot_result


# 遍历excel文件各个sheet,对“区域经理”为“省区合计”、“大区合计”的行进行着色
def highlight_summary_rows(file_name, sheng_fill_color="EEECE1", daqu_fill_color="948A54"):
    workbook = openpyxl.load_workbook(file_name)
    for worksheet in workbook.worksheets:
        quyu_column_index = None
        for column in worksheet.columns:
            if column[0].value == "区域经理":
                #quyu_column_index = column[0].column_letter  # 查找区域经理列
                quyu_column_index = column[0].column  # 查找区域经理列
                quyu_column_index = quyu_column_index -1
                break
    
        if quyu_column_index is not None:
            for row in worksheet.iter_rows():
                if row[quyu_column_index].value == "省区合计":
                    for cell in row[0:18]:
                        cell.fill = PatternFill(start_color=sheng_fill_color, end_color=sheng_fill_color, fill_type="solid")
                elif row[quyu_column_index].value == "大区合计":
                    for cell in row[0:18]:
                        cell.fill = PatternFill(start_color=daqu_fill_color, end_color=daqu_fill_color, fill_type="solid")
    
    workbook.save(file_name)

def set_percentage_format(file_name, patterns=None):
    if patterns is None:
        patterns = r'(占比|同比|环比|对比)'

    workbook = openpyxl.load_workbook(file_name)
    for worksheet in workbook.worksheets:
        for column in worksheet.columns:
            if re.search(patterns, column[0].value) and not re.search(r'期', column[0].value):
                for cell in column:
                    cell.number_format = '0.00%'

    workbook.save(file_name)
    

#添加各级经理汇总行
def add_summary(df):
    df_pivot = pd.pivot_table(
        df,
        index=["大区经理", "省区经理", "区域经理"],
        aggfunc="sum",
        fill_value=0,
    )
    summary_by_daqu_manager = df_pivot.groupby(level="大区经理").sum().reset_index()
    summary_by_sheng_manager = df_pivot.groupby(level=["大区经理", "省区经理"]).sum().reset_index()
    df_pivot = df_pivot.reset_index()
    result = pd.concat([df_pivot, summary_by_daqu_manager, summary_by_sheng_manager], axis=0)
    result = result.reset_index()
    
    result = result.sort_values(["大区经理", "省区经理", "区域经理"], ascending=True)

    # 查找“区域经理”列的空值并根据“省区经理”列是否为空进行修改
    result["区域经理"] = result.apply(
        lambda row: "大区合计"
        if pd.isna(row["区域经理"]) and pd.isna(row["省区经理"])
        else "省区合计"
        if pd.isna(row["区域经理"])
        else row["区域经理"],
        axis=1,
    )

    return result

# 列出文件夹中所有文件    
def list_files(folder):
    return [os.path.join(folder, f) for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]

# 定义一个函数，用于去除括号及其内容
def remove_brackets(text):
    return text.str.replace('(\(.*?\))', '', regex=True)